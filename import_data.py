#imports
import os
import django
from openpyxl import load_workbook
from datetime import datetime

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "portal.settings")
django.setup()

#import database tables 
from core.models import (TblDepartment, 
                         TblUser, 
                         TblTeam, 
                         TblProject,
                         TblDependencies)

Excel_File = "team_registry.xlsx"

#clean the data
def clean(value):
    #clean values from excel so no empty cells or extra spaces
    if value is None:
        return None
    value = str(value).strip()

    if value == "" or value == "#REF!":
        return None

    return value

#create users
def make_username(full_name):
    #changes a users full name into a username seperated by a .
    return full_name.lower().replace(" ",".")

#create user record for team leader/department leader
def get_or_create_user(full_name,role):

    full_name = clean(full_name)

    if not full_name:
        return None
    
    users_names=full_name.split()
    first_name = users_names[0]
    if len(users_names)>1:
        surname = users_names[-1]
    else:
        surname = ""
    
    username = make_username(full_name)
    email = f"{username}@sky.com"

    #checks to see if a suer exists with that username, if found returns it if not new user is created
    user,created = TblUser.objects.get_or_create(
        uname= username,
        defaults={
            "fname" : first_name,
            "sname": surname,
            "email": email,
            "password": "Password123",
            "role": role,
            "created" : datetime.now(),
            "active": True,
        }
    )

    return user

#converts excel values into ints
def get_integer(value):
    value = clean(value)

    if value is None:
        return None
    
    try:

        return int(float(value))
    except ValueError:
        return None
    
#read excel spreadsheet and creates database entries
def import_data():

    #open excel file
    workbook = load_workbook(Excel_File)
    sheet = workbook.active

    #First row in spreadsheet has column names
    headers = [cell.value for cell in sheet[1]]

    #set all counts to 0
    imported_departments = 0
    imported_users = 0
    imported_teams = 0
    imported_projects = 0
    imported_dependencies = 0

    #loop through each row starting from 2
    for row in sheet.iter_rows(min_row=2, values_only=True):
        #creates dictionary 
        row_data = dict(zip(headers,row))

        department_name = clean(row_data.get("Department"))
        team_leader_name = clean(row_data.get("Team Leader"))
        department_head_name = clean(row_data.get("Department Head"))
        team_name = clean(row_data.get("Team Name"))
        project_name = clean(row_data.get("Jira Project Name"))
        workstream = clean(row_data.get("Workstream (MF)"))
        repo_url = clean(row_data.get("Project (codebase) (Github Repo)"))
        jira_board = clean(row_data.get("Jira board Link"))
        focus_areas = clean(row_data.get("Development Focus Areas"))
        skills = clean(row_data.get("Key Skills & Technologies"))
        downstream_dependency = clean(row_data.get("Downstream Dependencies"))
        dependency_type = clean(row_data.get("Dependency Type"))
        software_owned = clean(row_data.get("Software Owned and Evolved By This Team"))
        versioning = clean(row_data.get("Versioning Approaches"))
        wiki_search_terms = clean(row_data.get("Wiki Search Terms"))
        slack = clean(row_data.get("Slack Channels"))
        daily_standup = clean(row_data.get("Daily Standup Time and Link"))
        agile_practices = clean(row_data.get("Agile Practices"))
        wiki = clean(row_data.get("Team Wiki"))
        active_projects = get_integer(row_data.get(" # of Concurrent Projects"))

        if not team_name:
            continue
        
        #add department heads
        department_head = get_or_create_user(department_head_name, "Department Head")
        team_leader = get_or_create_user(team_leader_name, "Team Leader")

        if department_head:
            imported_users+=1
        
        if team_leader:
            imported_users+=1

        department,department_created = TblDepartment.objects.get_or_create(
            name=department_name,
            defaults={
                "department_head": department_head,
            }
        )

        if department_created:
            imported_departments+=1

        team,team_created = TblTeam.objects.get_or_create(
            name=team_name,
            defaults={
                "team_leader":team_leader,
                "department" : department,
                "description": focus_areas,
                "skills_and_tech":skills,
                "software_owned_and_evolved": software_owned,
                "versioning": versioning,
                "agile_practices":agile_practices,
                "slack": slack,
                "wiki": wiki,
                "wiki_search_terms":wiki_search_terms,
                "daily_standup": None,
                "active_projects": active_projects,
            }
        )

        if team_created:
            imported_teams+=1

        if project_name:
            project,project_created = TblProject.objects.get_or_create(
                name=project_name,
                team=team,
                defaults={
                    "description": workstream,
                    "codebase": repo_url,
                    "jira_board": jira_board,
                    "created": datetime.now(),
                    "status": "Active",
                }
                
            )

            if project_created:
                imported_projects+=1

        
        
        if downstream_dependency:
            dependency,dependency_created = TblDependencies.objects.get_or_create(
                team=team,
                type= dependency_type,
                defaults={
                    "upstream": False,
                    "downstream": True,
                }
            )
        
            if dependency_created:
                imported_dependencies+=1

    print("Import complete.")
    print("Departments imported:", imported_departments)
    print("Users processed:", imported_users)
    print("Teams imported:", imported_teams)
    print("Projects imported:", imported_projects)
    print("Dependencies imported:", imported_dependencies)
    print("Total departments:", TblDepartment.objects.count())
    print("Total users:", TblUser.objects.count())
    print("Total teams:", TblTeam.objects.count())
    print("Total projects:", TblProject.objects.count())
    print("Total dependencies:", TblDependencies.objects.count())

import_data()
